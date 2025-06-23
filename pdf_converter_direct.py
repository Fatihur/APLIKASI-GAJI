"""
PDF Converter Direct Method
Konversi Excel ke PDF tanpa membuka Excel application
"""

import os
import pandas as pd
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
from watermark_manager import WatermarkManager

class PDFConverterDirect:
    def __init__(self, enable_watermark=True, watermark_opacity=0.3, watermark_position="bottom-right"):
        """Initialize direct PDF converter"""
        self.styles = getSampleStyleSheet()
        self.enable_watermark = enable_watermark
        self.watermark_manager = WatermarkManager() if enable_watermark else None
        self.watermark_opacity = watermark_opacity
        self.watermark_position = watermark_position
        
    def convert_excel_to_pdf_direct(self, excel_file, selected_sheets, output_directory, folder_prefix=""):
        """
        Konversi Excel ke PDF tanpa membuka Excel application
        
        Args:
            excel_file (str): Path ke file Excel
            selected_sheets (list): List nama sheet yang akan dikonversi
            output_directory (str): Direktori output
            folder_prefix (str): Prefix untuk nama file
            
        Returns:
            dict: Dictionary hasil konversi {sheet_name: pdf_path}
        """
        results = {}
        
        # Buat direktori output jika belum ada
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
        
        try:
            # Baca Excel file menggunakan openpyxl (tidak membuka Excel)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            for sheet_name in selected_sheets:
                try:
                    if sheet_name not in workbook.sheetnames:
                        print(f"Sheet '{sheet_name}' not found in workbook")
                        results[sheet_name] = None
                        continue
                    
                    # Nama file output dengan prefix
                    safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
                    if folder_prefix:
                        pdf_filename = f"{folder_prefix}_{safe_sheet_name}.pdf"
                    else:
                        base_name = os.path.splitext(os.path.basename(excel_file))[0]
                        pdf_filename = f"{base_name}_{safe_sheet_name}.pdf"
                    
                    pdf_path = os.path.join(output_directory, pdf_filename)
                    
                    # Konversi sheet ke PDF
                    success = self._convert_sheet_to_pdf(workbook, sheet_name, pdf_path)

                    if success:
                        # Tambahkan watermark jika enabled
                        if self.enable_watermark and self.watermark_manager:
                            watermark_success = self.watermark_manager.add_watermark_to_pdf(
                                pdf_path,
                                opacity=self.watermark_opacity,
                                position=self.watermark_position
                            )
                            if not watermark_success:
                                print(f"⚠️  Failed to add watermark to {sheet_name}")

                        results[sheet_name] = pdf_path
                    else:
                        results[sheet_name] = None
                        
                except Exception as e:
                    print(f"Error converting sheet '{sheet_name}': {str(e)}")
                    results[sheet_name] = None
            
            workbook.close()
            
        except Exception as e:
            print(f"Error opening Excel file: {str(e)}")
            
        return results
    
    def _convert_sheet_to_pdf(self, workbook, sheet_name, output_path):
        """
        Konversi single sheet ke PDF
        
        Args:
            workbook: Openpyxl workbook object
            sheet_name (str): Nama sheet
            output_path (str): Path output PDF
            
        Returns:
            bool: True jika berhasil
        """
        try:
            worksheet = workbook[sheet_name]
            
            # Dapatkan data dari worksheet
            data, formatting = self._extract_sheet_data(worksheet)
            
            if not data:
                print(f"No data found in sheet '{sheet_name}'")
                return False
            
            # Tentukan orientasi berdasarkan jumlah kolom
            num_cols = len(data[0]) if data else 0
            use_landscape = num_cols > 6
            page_size = landscape(A4) if use_landscape else A4
            
            # Buat dokumen PDF
            doc = SimpleDocTemplate(
                output_path,
                pagesize=page_size,
                rightMargin=15*mm,
                leftMargin=15*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )
            
            elements = []
            
            # Tambahkan judul
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Heading1'],
                fontSize=14,
                spaceAfter=10,
                alignment=1,  # Center alignment
                textColor=colors.darkblue
            )
            title = Paragraph(f"<b>{sheet_name}</b>", title_style)
            elements.append(title)
            elements.append(Spacer(1, 8))
            
            # Buat tabel
            table = Table(data)
            
            # Apply styling
            table_style = self._create_table_style(data, formatting)
            table.setStyle(table_style)
            
            # Atur lebar kolom
            if num_cols > 0:
                available_width = page_size[0] - 30*mm
                col_width = available_width / num_cols
                table._argW = [col_width] * num_cols
            
            elements.append(table)
            
            # Tambahkan watermark jika enabled
            if self.enable_watermark and self.watermark_manager and self.watermark_manager.watermark_exists:
                watermark_element = self._create_watermark_element(page_size, self.watermark_position, self.watermark_opacity)
                if watermark_element:
                    elements.append(watermark_element)

            # Build PDF
            doc.build(elements)

            return True
            
        except Exception as e:
            print(f"Error creating PDF for sheet '{sheet_name}': {str(e)}")
            return False
    
    def _extract_sheet_data(self, worksheet):
        """
        Extract data dan formatting dari worksheet
        
        Args:
            worksheet: Openpyxl worksheet object
            
        Returns:
            tuple: (data, formatting)
        """
        data = []
        formatting = {}
        
        # Dapatkan range yang berisi data
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row == 1 and max_col == 1:
            # Check if the single cell is empty
            cell_value = worksheet.cell(1, 1).value
            if cell_value is None:
                return [], {}
        
        # Extract data dan formatting
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                
                # Get cell value
                cell_value = cell.value
                if cell_value is None:
                    cell_value = ""
                else:
                    cell_value = str(cell_value)
                
                row_data.append(cell_value)
                
                # Get cell formatting
                cell_coord = f"{row_idx}_{col_idx}"
                formatting[cell_coord] = {
                    'font': cell.font,
                    'fill': cell.fill,
                    'alignment': cell.alignment
                }
            
            # Skip completely empty rows at the end
            if any(cell.strip() for cell in row_data if isinstance(cell, str)):
                data.append(row_data)
            elif row_idx <= 10:  # Keep first 10 rows even if empty
                data.append(row_data)
        
        return data, formatting
    
    def _create_table_style(self, data, formatting):
        """
        Buat style untuk tabel berdasarkan Excel formatting
        
        Args:
            data (list): Data tabel
            formatting (dict): Formatting information
            
        Returns:
            TableStyle: Style untuk tabel
        """
        style_commands = [
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            
            # Grid and borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.Color(0.2, 0.4, 0.6)),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]
        
        # Add alternating row colors
        if len(data) > 1:
            for row_idx in range(1, len(data)):
                if row_idx % 2 == 0:
                    style_commands.append(
                        ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.Color(0.95, 0.95, 0.95))
                    )
        
        # Apply Excel formatting if available
        if formatting:
            style_commands.extend(self._apply_excel_formatting(data, formatting))
        
        return TableStyle(style_commands)
    
    def _apply_excel_formatting(self, data, formatting):
        """
        Apply Excel formatting ke table style
        
        Args:
            data (list): Data tabel
            formatting (dict): Formatting information
            
        Returns:
            list: Additional style commands
        """
        additional_styles = []
        
        for row_idx, row in enumerate(data):
            for col_idx, cell_value in enumerate(row):
                cell_coord = f"{row_idx + 1}_{col_idx + 1}"
                
                if cell_coord in formatting:
                    cell_format = formatting[cell_coord]
                    
                    # Apply font formatting
                    if cell_format['font'] and cell_format['font'].bold:
                        additional_styles.append(
                            ('FONTNAME', (col_idx, row_idx), (col_idx, row_idx), 'Helvetica-Bold')
                        )
                    
                    # Apply background color
                    if cell_format['fill'] and cell_format['fill'].start_color:
                        try:
                            color_hex = cell_format['fill'].start_color.rgb
                            if color_hex and color_hex != '00000000':
                                # Convert hex to RGB
                                if len(color_hex) == 8:  # ARGB format
                                    color_hex = color_hex[2:]  # Remove alpha
                                
                                r = int(color_hex[0:2], 16) / 255.0
                                g = int(color_hex[2:4], 16) / 255.0
                                b = int(color_hex[4:6], 16) / 255.0
                                
                                additional_styles.append(
                                    ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), 
                                     colors.Color(r, g, b))
                                )
                        except:
                            pass  # Skip if color conversion fails
                    
                    # Apply text alignment
                    if cell_format['alignment']:
                        alignment = cell_format['alignment'].horizontal
                        if alignment == 'center':
                            additional_styles.append(
                                ('ALIGN', (col_idx, row_idx), (col_idx, row_idx), 'CENTER')
                            )
                        elif alignment == 'right':
                            additional_styles.append(
                                ('ALIGN', (col_idx, row_idx), (col_idx, row_idx), 'RIGHT')
                            )
        
        return additional_styles

    def _create_watermark_element(self, page_size, position, opacity):
        """
        Buat watermark element untuk ditambahkan ke PDF

        Args:
            page_size: Ukuran halaman
            position (str): Posisi watermark
            opacity (float): Transparansi watermark

        Returns:
            Spacer atau Image element untuk watermark
        """
        try:
            if not self.watermark_manager or not self.watermark_manager.watermark_exists:
                return None

            from reportlab.platypus import Image as RLImage
            from reportlab.lib.units import mm

            # Load watermark image
            watermark_path = self.watermark_manager.watermark_path

            # Tentukan ukuran watermark (maksimal 30% dari halaman)
            page_width, page_height = page_size
            max_width = page_width * 0.3
            max_height = page_height * 0.3

            # Buat watermark image element
            watermark_img = RLImage(
                watermark_path,
                width=max_width,
                height=max_height
            )

            # Set transparency (simplified - reportlab doesn't support direct opacity)
            # Watermark akan ditampilkan sebagai image biasa

            return watermark_img

        except Exception as e:
            print(f"❌ Error creating watermark element: {str(e)}")
            return None
    
    def convert_single_sheet_direct(self, excel_file, sheet_name, output_path, folder_prefix=""):
        """
        Konversi single sheet ke PDF tanpa membuka Excel
        
        Args:
            excel_file (str): Path ke file Excel
            sheet_name (str): Nama sheet
            output_path (str): Path output PDF
            folder_prefix (str): Prefix untuk nama file
            
        Returns:
            bool: True jika berhasil
        """
        try:
            # Buat direktori output jika belum ada
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # Load workbook
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            # Convert sheet
            success = self._convert_sheet_to_pdf(workbook, sheet_name, output_path)
            
            workbook.close()
            
            return success
            
        except Exception as e:
            print(f"Error converting sheet '{sheet_name}': {str(e)}")
            return False
