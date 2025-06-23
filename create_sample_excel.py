"""
Script untuk membuat file Excel contoh untuk testing
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_sample_excel():
    """Membuat file Excel contoh dengan multiple sheets"""
    
    # Buat workbook baru
    wb = openpyxl.Workbook()
    
    # Hapus sheet default
    wb.remove(wb.active)
    
    # Sheet 1: Data Karyawan
    ws1 = wb.create_sheet("Data Karyawan")
    
    # Header
    headers1 = ["ID", "Nama", "Jabatan", "Gaji", "Departemen"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Data karyawan
    data1 = [
        [1, "Ahmad Rizki", "Manager", 15000000, "IT"],
        [2, "Siti Nurhaliza", "Developer", 12000000, "IT"],
        [3, "Budi Santoso", "Analyst", 10000000, "Finance"],
        [4, "Maya Sari", "Designer", 9000000, "Marketing"],
        [5, "Andi Wijaya", "Admin", 7000000, "HR"],
        [6, "Dewi Lestari", "Tester", 8500000, "IT"],
        [7, "Rudi Hartono", "Supervisor", 11000000, "Production"],
        [8, "Lina Marlina", "Coordinator", 9500000, "Marketing"]
    ]
    
    for row, data_row in enumerate(data1, 2):
        for col, value in enumerate(data_row, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if col == 4:  # Kolom gaji
                cell.number_format = '#,##0'
    
    # Atur lebar kolom
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    
    # Sheet 2: Laporan Penjualan
    ws2 = wb.create_sheet("Laporan Penjualan")
    
    # Header
    headers2 = ["Bulan", "Produk A", "Produk B", "Produk C", "Total"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Data penjualan
    data2 = [
        ["Januari", 1500000, 2000000, 1200000, "=B2+C2+D2"],
        ["Februari", 1800000, 2200000, 1400000, "=B3+C3+D3"],
        ["Maret", 2000000, 2500000, 1600000, "=B4+C4+D4"],
        ["April", 1700000, 2100000, 1300000, "=B5+C5+D5"],
        ["Mei", 2200000, 2800000, 1800000, "=B6+C6+D6"],
        ["Juni", 2500000, 3000000, 2000000, "=B7+C7+D7"]
    ]
    
    for row, data_row in enumerate(data2, 2):
        for col, value in enumerate(data_row, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if col > 1:  # Kolom angka
                cell.number_format = '#,##0'
    
    # Atur lebar kolom
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    
    # Judul
    ws3.cell(row=1, column=1, value="RINGKASAN LAPORAN").font = Font(bold=True, size=16)
    ws3.merge_cells('A1:D1')
    ws3['A1'].alignment = Alignment(horizontal="center")
    
    # Info
    info_data = [
        ["Total Karyawan:", 8],
        ["Rata-rata Gaji:", "=AVERAGE('Data Karyawan'!D2:D9)"],
        ["Total Penjualan Q1:", "=SUM('Laporan Penjualan'!E2:E4)"],
        ["Total Penjualan Q2:", "=SUM('Laporan Penjualan'!E5:E7)"]
    ]
    
    for row, data_row in enumerate(info_data, 3):
        ws3.cell(row=row, column=1, value=data_row[0]).font = Font(bold=True)
        cell = ws3.cell(row=row, column=2, value=data_row[1])
        if isinstance(data_row[1], str) and data_row[1].startswith('='):
            cell.number_format = '#,##0'
    
    # Atur lebar kolom
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20
    
    # Simpan file
    wb.save("sample_data.xlsx")
    print("File sample_data.xlsx berhasil dibuat!")

if __name__ == "__main__":
    create_sample_excel()
