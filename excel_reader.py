"""
Excel Reader Module
Modul untuk membaca file Excel dan mengekstrak informasi sheet
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

class ExcelReader:
    def __init__(self, file_path):
        """
        Initialize Excel reader dengan file path
        
        Args:
            file_path (str): Path ke file Excel
        """
        self.file_path = file_path
        self.workbook = None
        self.load_workbook()
        
    def load_workbook(self):
        """Load workbook dari file Excel"""
        try:
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"File tidak ditemukan: {self.file_path}")
                
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
        except Exception as e:
            raise Exception(f"Error loading Excel file: {str(e)}")
    
    def get_sheets_info(self):
        """
        Mendapatkan informasi semua sheet dalam workbook
        
        Returns:
            dict: Dictionary dengan nama sheet sebagai key dan info sheet sebagai value
        """
        if not self.workbook:
            raise Exception("Workbook belum di-load")
            
        sheets_info = {}
        
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            
            # Hitung jumlah baris dan kolom yang terisi
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Hitung sel yang terisi
            filled_cells = 0
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, 
                                         min_col=1, max_col=max_col):
                for cell in row:
                    if cell.value is not None:
                        filled_cells += 1
            
            sheets_info[sheet_name] = {
                'max_row': max_row,
                'max_col': max_col,
                'filled_cells': filled_cells,
                'worksheet': worksheet
            }
            
        return sheets_info
    
    def get_sheet_data(self, sheet_name):
        """
        Mendapatkan data dari sheet tertentu
        
        Args:
            sheet_name (str): Nama sheet
            
        Returns:
            list: List of lists berisi data sheet
        """
        if not self.workbook:
            raise Exception("Workbook belum di-load")
            
        if sheet_name not in self.workbook.sheetnames:
            raise Exception(f"Sheet '{sheet_name}' tidak ditemukan")
            
        worksheet = self.workbook[sheet_name]
        data = []
        
        for row in worksheet.iter_rows(values_only=True):
            # Konversi None ke string kosong untuk konsistensi
            row_data = [str(cell) if cell is not None else "" for cell in row]
            data.append(row_data)
            
        return data
    
    def get_sheet_with_formatting(self, sheet_name):
        """
        Mendapatkan sheet dengan informasi formatting
        
        Args:
            sheet_name (str): Nama sheet
            
        Returns:
            dict: Dictionary berisi data dan formatting info
        """
        if not self.workbook:
            raise Exception("Workbook belum di-load")
            
        if sheet_name not in self.workbook.sheetnames:
            raise Exception(f"Sheet '{sheet_name}' tidak ditemukan")
            
        worksheet = self.workbook[sheet_name]
        
        # Data dengan formatting
        formatted_data = {
            'data': [],
            'formatting': {},
            'merged_cells': [],
            'column_widths': {},
            'row_heights': {}
        }
        
        # Ambil data dan formatting
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            row_data = []
            for col_idx, cell in enumerate(row, 1):
                cell_value = str(cell.value) if cell.value is not None else ""
                row_data.append(cell_value)
                
                # Simpan formatting info
                cell_coord = f"{get_column_letter(col_idx)}{row_idx}"
                formatted_data['formatting'][cell_coord] = {
                    'font_bold': cell.font.bold if cell.font else False,
                    'font_size': cell.font.size if cell.font else 11,
                    'font_color': str(cell.font.color.rgb) if cell.font and cell.font.color else None,
                    'fill_color': str(cell.fill.start_color.rgb) if cell.fill and cell.fill.start_color else None,
                    'alignment': {
                        'horizontal': cell.alignment.horizontal if cell.alignment else None,
                        'vertical': cell.alignment.vertical if cell.alignment else None
                    },
                    'border': bool(cell.border.left.style or cell.border.right.style or 
                                 cell.border.top.style or cell.border.bottom.style) if cell.border else False
                }
                
            formatted_data['data'].append(row_data)
        
        # Ambil merged cells
        for merged_range in worksheet.merged_cells.ranges:
            formatted_data['merged_cells'].append(str(merged_range))
        
        # Ambil column widths
        for col_letter, col_dimension in worksheet.column_dimensions.items():
            if col_dimension.width:
                formatted_data['column_widths'][col_letter] = col_dimension.width
        
        # Ambil row heights
        for row_num, row_dimension in worksheet.row_dimensions.items():
            if row_dimension.height:
                formatted_data['row_heights'][row_num] = row_dimension.height
                
        return formatted_data
    
    def get_sheet_names(self):
        """
        Mendapatkan daftar nama sheet
        
        Returns:
            list: List nama sheet
        """
        if not self.workbook:
            raise Exception("Workbook belum di-load")
            
        return self.workbook.sheetnames
    
    def close(self):
        """Tutup workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
