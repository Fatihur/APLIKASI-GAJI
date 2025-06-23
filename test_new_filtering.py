"""
Script untuk test new filtering system
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import sys

def create_test_excel_with_ignored_sheets():
    """Create test Excel file with sheets that should be ignored"""
    print("📁 Creating test Excel with ignored sheets...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Add sheets that should be ignored
    ignored_sheets = [
        'Payroll adjust',
        'Database', 
        'Summary Amman',
        'Summary Karyawan',
        'PPh 21',
        'Payroll',
        'Tarif TER',
        'hr_libur',
        'jm_istrht'
    ]
    
    # Add sheets that should NOT be ignored
    valid_sheets = [
        'Data Karyawan',
        'Laporan Bulanan',
        'Absensi',
        'Overtime Report',
        'Employee List'
    ]
    
    # Create ignored sheets
    for sheet_name in ignored_sheets:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = f"This is {sheet_name} sheet"
        ws['A1'].font = Font(bold=True)
        ws['A2'] = "This sheet should be ignored"
    
    # Create valid sheets with some data
    for sheet_name in valid_sheets:
        ws = wb.create_sheet(sheet_name)
        
        # Add header
        headers = ['ID', 'Name', 'Department', 'Position']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add sample data
        sample_data = [
            [1, 'John Doe', 'IT', 'Developer'],
            [2, 'Jane Smith', 'HR', 'Manager'],
            [3, 'Bob Johnson', 'Finance', 'Analyst']
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save file
    filename = "test_filtering.xlsx"
    wb.save(filename)
    print(f"   ✅ Created: {filename}")
    print(f"   📋 Ignored sheets: {len(ignored_sheets)}")
    print(f"   📋 Valid sheets: {len(valid_sheets)}")
    
    return filename, ignored_sheets, valid_sheets

def test_sheet_filtering():
    """Test the new sheet filtering logic"""
    print("\n🔍 Testing New Sheet Filtering Logic...")
    print("=" * 50)
    
    # Import the filtering function
    sys.path.append('.')
    from main import ExcelToPDFApp
    
    # Create a dummy app to test filtering
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()  # Hide window
    
    app = ExcelToPDFApp(root)
    
    # Test cases with new ignored sheets
    test_cases = [
        # Should be ignored
        ('Payroll adjust', True),
        ('Database', True),
        ('Summary Amman', True),
        ('Summary Karyawan', True),
        ('PPh 21', True),
        ('Payroll', True),
        ('Tarif TER', True),
        ('hr_libur', True),
        ('jm_istrht', True),
        
        # Case variations (should still be ignored)
        ('payroll adjust', True),  # lowercase
        ('DATABASE', True),        # uppercase
        ('Summary amman', True),   # mixed case
        
        # Should NOT be ignored
        ('Data Karyawan', False),
        ('Laporan Bulanan', False),
        ('Absensi', False),
        ('Overtime Report', False),
        ('Employee List', False),
        ('Sheet1', False),
        ('Summary Report', False),  # Different from "Summary Amman"
        ('Payroll Report', False),  # Different from "Payroll"
    ]
    
    print("Testing sheet name filtering:")
    passed = 0
    failed_cases = []
    
    for sheet_name, should_ignore in test_cases:
        result = app.is_sheet_ignored(sheet_name)
        status = "✅" if result == should_ignore else "❌"
        print(f"   {status} '{sheet_name}' -> Ignored: {result} (Expected: {should_ignore})")
        
        if result == should_ignore:
            passed += 1
        else:
            failed_cases.append((sheet_name, result, should_ignore))
    
    root.destroy()
    
    print(f"\nSheet filtering test: {passed}/{len(test_cases)} passed")
    
    if failed_cases:
        print("\n❌ Failed cases:")
        for sheet_name, got, expected in failed_cases:
            print(f"   '{sheet_name}': got {got}, expected {expected}")
    
    return passed == len(test_cases)

def test_excel_file_with_filtering():
    """Test filtering with actual Excel file"""
    print("\n📄 Testing Excel File with Filtering...")
    print("=" * 50)
    
    # Create test file
    filename, ignored_sheets, valid_sheets = create_test_excel_with_ignored_sheets()
    
    try:
        # Test with ExcelReader
        from excel_reader import ExcelReader
        
        reader = ExcelReader(filename)
        all_sheets = reader.get_sheet_names()
        
        print(f"📋 Total sheets in file: {len(all_sheets)}")
        print(f"   All sheets: {all_sheets}")
        
        # Test filtering
        sys.path.append('.')
        from main import ExcelToPDFApp
        
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        
        app = ExcelToPDFApp(root)
        
        filtered_sheets = []
        ignored_count = 0
        
        for sheet_name in all_sheets:
            if app.is_sheet_ignored(sheet_name):
                ignored_count += 1
                print(f"   🚫 Ignored: {sheet_name}")
            else:
                filtered_sheets.append(sheet_name)
                print(f"   ✅ Included: {sheet_name}")
        
        root.destroy()
        reader.close()
        
        print(f"\n📊 Filtering Results:")
        print(f"   Total sheets: {len(all_sheets)}")
        print(f"   Ignored: {ignored_count}")
        print(f"   Included: {len(filtered_sheets)}")
        print(f"   Included sheets: {filtered_sheets}")
        
        # Verify results
        expected_ignored = len(ignored_sheets)
        expected_included = len(valid_sheets)
        
        success = (ignored_count == expected_ignored and 
                  len(filtered_sheets) == expected_included)
        
        if success:
            print("   ✅ Filtering working correctly!")
        else:
            print(f"   ❌ Expected {expected_ignored} ignored, {expected_included} included")
        
        return success
        
    except Exception as e:
        print(f"❌ Error testing Excel file: {str(e)}")
        return False
        
    finally:
        # Cleanup
        if os.path.exists(filename):
            os.remove(filename)
            print(f"   🧹 Cleaned up: {filename}")

def main():
    """Main test function"""
    print("🧪 New Filtering System Test Suite")
    print("=" * 60)
    
    tests = [
        ("Sheet Filtering Logic", test_sheet_filtering),
        ("Excel File Filtering", test_excel_file_with_filtering)
    ]
    
    results = []
    
    for test_name, test_func in tests:
        print(f"\n🔬 Running {test_name} test...")
        try:
            result = test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"❌ {test_name} test crashed: {str(e)}")
            results.append((test_name, False))
    
    # Summary
    print("\n" + "=" * 60)
    print("📋 TEST SUMMARY")
    print("=" * 60)
    
    passed = 0
    for test_name, result in results:
        status = "✅ PASSED" if result else "❌ FAILED"
        print(f"{test_name:.<30} {status}")
        if result:
            passed += 1
    
    print(f"\nOverall: {passed}/{len(results)} tests passed")
    
    if passed == len(results):
        print("🎉 All filtering tests passed!")
        print("\n📋 New Ignored Sheets:")
        ignored_list = [
            'Payroll adjust', 'Database', 'Summary Amman', 'Summary Karyawan',
            'PPh 21', 'Payroll', 'Tarif TER', 'hr_libur', 'jm_istrht'
        ]
        for sheet in ignored_list:
            print(f"   🚫 {sheet}")
    else:
        print("⚠️  Some tests failed. Check the output above for details.")

if __name__ == "__main__":
    main()
